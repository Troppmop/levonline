import io

import openpyxl
import pytest
from httpx import AsyncClient


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.mark.asyncio
async def test_list_data_tables_is_admin_only(client: AsyncClient, auth_headers: dict[str, str]):
    resp = await client.get("/api/v1/admin/data-tables", headers=auth_headers)
    assert resp.status_code == 200
    table_names = [t["table_name"] for t in resp.json()]
    assert "residents" in table_names
    assert "users" in table_names

    unauth = await client.get("/api/v1/admin/data-tables")
    assert unauth.status_code == 401


@pytest.mark.asyncio
async def test_staff_cannot_reach_admin_data(client: AsyncClient, make_login):
    from app.models.enums import UserRole

    headers = await make_login("staffer@lev.org", "password123", UserRole.STAFF)
    resp = await client.get("/api/v1/admin/data-tables", headers=headers)
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_table_returns_xlsx(client: AsyncClient, auth_headers: dict[str, str]):
    await client.post(
        "/api/v1/residents", json={"first_name": "Export", "last_name": "Test"}, headers=auth_headers
    )
    resp = await client.get("/api/v1/admin/export/residents", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "first_name" in header_row
    assert "id" in header_row


@pytest.mark.asyncio
async def test_export_all_has_one_sheet_per_table(client: AsyncClient, auth_headers: dict[str, str]):
    resp = await client.get("/api/v1/admin/export-all", headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "residents" in wb.sheetnames
    assert "users" in wb.sheetnames
    assert "hashed_password" not in [c.value for c in next(wb["users"].iter_rows(min_row=1, max_row=1))]


@pytest.mark.asyncio
async def test_import_creates_and_updates_rooms(client: AsyncClient, auth_headers: dict[str, str]):
    create_content = _xlsx_bytes(["floor", "room_number", "capacity"], [[1, "101", 2]])
    resp = await client.post(
        "/api/v1/admin/import/rooms",
        files={"file": ("rooms.xlsx", create_content, "application/octet-stream")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["inserted_count"] == 1
    assert body["updated_count"] == 0
    assert body["errors"] == []

    rooms_resp = await client.get("/api/v1/rooms", headers=auth_headers)
    room = next(r for r in rooms_resp.json() if r["room_number"] == "101")
    assert room["capacity"] == 2

    # Re-importing the same (floor, room_number) updates instead of duplicating.
    update_content = _xlsx_bytes(["floor", "room_number", "capacity"], [[1, "101", 5]])
    resp2 = await client.post(
        "/api/v1/admin/import/rooms",
        files={"file": ("rooms.xlsx", update_content, "application/octet-stream")},
        headers=auth_headers,
    )
    body2 = resp2.json()
    assert body2["inserted_count"] == 0
    assert body2["updated_count"] == 1

    rooms_resp2 = await client.get("/api/v1/rooms", headers=auth_headers)
    matching = [r for r in rooms_resp2.json() if r["room_number"] == "101"]
    assert len(matching) == 1
    assert matching[0]["capacity"] == 5


@pytest.mark.asyncio
async def test_import_dry_run_does_not_write(client: AsyncClient, auth_headers: dict[str, str]):
    content = _xlsx_bytes(["floor", "room_number", "capacity"], [[4, "401", 3]])
    resp = await client.post(
        "/api/v1/admin/import/rooms?dry_run=true",
        files={"file": ("rooms.xlsx", content, "application/octet-stream")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["dry_run"] is True
    assert body["inserted_count"] == 1

    rooms_resp = await client.get("/api/v1/rooms", headers=auth_headers)
    assert not any(r["room_number"] == "401" for r in rooms_resp.json())


@pytest.mark.asyncio
async def test_import_is_all_or_nothing_on_bad_row(client: AsyncClient, auth_headers: dict[str, str]):
    content = _xlsx_bytes(
        ["floor", "room_number", "capacity"],
        [[1, "201", 2], ["not-a-number", "202", 2]],
    )
    resp = await client.post(
        "/api/v1/admin/import/rooms",
        files={"file": ("rooms.xlsx", content, "application/octet-stream")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["inserted_count"] == 0
    assert body["updated_count"] == 0
    assert len(body["errors"]) == 1
    assert body["errors"][0]["row"] == 3

    rooms_resp = await client.get("/api/v1/rooms", headers=auth_headers)
    assert not any(r["room_number"] == "201" for r in rooms_resp.json())


@pytest.mark.asyncio
async def test_import_cannot_create_new_users(client: AsyncClient, auth_headers: dict[str, str]):
    content = _xlsx_bytes(
        ["email", "full_name", "role", "is_active"],
        [["brandnew@lev.org", "Brand New", "staff", True]],
    )
    resp = await client.post(
        "/api/v1/admin/import/users",
        files={"file": ("users.xlsx", content, "application/octet-stream")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["inserted_count"] == 0
    assert len(body["errors"]) == 1
    assert "isn't supported via import" in body["errors"][0]["message"]


@pytest.mark.asyncio
async def test_import_rejects_unknown_columns(client: AsyncClient, auth_headers: dict[str, str]):
    content = _xlsx_bytes(["floor", "room_number", "not_a_real_column"], [[1, "301", "x"]])
    resp = await client.post(
        "/api/v1/admin/import/rooms",
        files={"file": ("rooms.xlsx", content, "application/octet-stream")},
        headers=auth_headers,
    )
    body = resp.json()
    assert any("not_a_real_column" in (e["column"] or "") for e in body["errors"])


@pytest.mark.asyncio
async def test_table_schema_describes_columns(client: AsyncClient, auth_headers: dict[str, str]):
    resp = await client.get("/api/v1/admin/schema/residents", headers=auth_headers)
    assert resp.status_code == 200
    body = resp.json()
    columns = {c["name"]: c for c in body["columns"]}
    assert columns["first_name"]["required"] is True
    assert columns["first_name"]["type"] == "text"
    assert columns["id"]["writable"] is False
    assert columns["status"]["type"] == "choice"
    assert set(columns["status"]["choices"]) == {"home", "away"}

    users_resp = await client.get("/api/v1/admin/schema/users", headers=auth_headers)
    users_columns = {c["name"] for c in users_resp.json()["columns"]}
    assert "hashed_password" not in users_columns


@pytest.mark.asyncio
async def test_import_template_download(client: AsyncClient, auth_headers: dict[str, str]):
    resp = await client.get("/api/v1/admin/import-template/residents", headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    header_row = [c.value for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
    assert "id" not in header_row
    assert "created_at" not in header_row
    assert "first_name" in header_row
